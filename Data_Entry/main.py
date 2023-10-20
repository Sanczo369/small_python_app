from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
import pathlib

root=Tk()
root.title("Data Entry")
root.geometry('700x400+300+200')
root.resizable(False, False)
root.configure(bg="#326273")
root.iconbitmap('logo.ico')

file=pathlib.Path('Backened_Data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Nazwa"
    sheet['B1']="Numer Telefonu"
    sheet['C1']="Wiek"
    sheet['D1']='Płeć'
    sheet['E1']='Adres'

    file.save('Backened_Data.xlsx')

def clear():
    name_value.set('')
    contact_value.set('')
    age_value.set('')
    address_entry.delete(1.0,END)


def submit():
    name=name_value.get()
    contact=contact_value.get()
    age=age_value.get()
    gender=gender_combobox.get()
    address=address_entry.get(1.0,END)
    file=openpyxl.load_workbook('Backened_Data.xlsx')
    sheet=file.active
    sheet.cell(column=1, row=sheet.max_row+1, value=name)
    sheet.cell(column=2, row=sheet.max_row, value=contact)
    sheet.cell(column=3, row=sheet.max_row, value=age)
    sheet.cell(column=4, row=sheet.max_row, value=gender)
    sheet.cell(column=5, row=sheet.max_row, value=address)
    file.save(r'Backened_Data.xlsx')

    messagebox.showinfo("info","Dodano!")
    name_value.set('')
    contact_value.set('')
    age_value.set('')
    address_entry.delete(1.0,END)


Label(root, text="Proszę o wypełnienie formularza zgłoszeniowego", font="arial 13", bg="#326273", fg="#fff").place(x=20,y=20)

Label(root, text="Imię", font=23, bg="#326273", fg="#fff").place(x=50,y=100)
Label(root, text="Numer Kontaktowy", font=23, bg="#326273", fg="#fff").place(x=50,y=150)
Label(root, text="Wiek", font=23, bg="#326273", fg="#fff").place(x=50,y=200)
Label(root, text="Płeć", font=23, bg="#326273", fg="#fff").place(x=370,y=200)
Label(root, text="Address", font=23, bg="#326273", fg="#fff").place(x=50,y=250)

name_value = StringVar()
contact_value = StringVar()
age_value = StringVar()

name_entry = Entry(root,textvariable=name_value, width=45, bd=2, font=20)
contact_entry = Entry(root,textvariable=contact_value, width=45, bd=2, font=20)
age_entry = Entry(root,textvariable=age_value, width=15, bd=2, font=20)
gender_combobox= Combobox(root,values=['M','K'], font="arial 14", state='r', width=14)
address_entry= Text(root, width=50, height=4, bd=4)

name_entry.place(x=200,y=100)
contact_entry.place(x=200,y=150)
age_entry.place(x=200,y=200)
gender_combobox.place(x=440,y=200)
address_entry.place(x=200,y=250)

Button(root, text="Zatwierd", bg="#326273", fg="white", width=15, height=2, command=submit).place(x=200,y=350)
Button(root, text="Wyczyść", bg="#326273", fg="white", width=15, height=2, command=clear).place(x=340,y=350)
Button(root, text="Zamknij", bg="#326273", fg="white", width=15, height=2, command=quit).place(x=480,y=350)

root.mainloop()